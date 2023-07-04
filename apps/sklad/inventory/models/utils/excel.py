from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.http import HttpResponse
from django.db.models import Prefetch
from apps.sklad.inventory.models.inventory import Inventory, InventoryItem, InventoryHistory
from django.template.defaultfilters import slugify


def export_inventory(request):
    # Get all inventory items with related inventory and history using Prefetch
    inventory_items = InventoryItem.objects.select_related('inventory').prefetch_related(
        Prefetch('inventory__history_entries', queryset=InventoryHistory.objects.order_by('-date_modified')))

    # Create a new workbook and get the active worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Define the column headers
    headers = ['Inventory ID', 'Материал', 'Количество', 'Описание', 'Дата добавления', 'Основание', 'Пользователь',
               'История - основание', 'История - дата']

    # Write the headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='center')

    # Write the inventory data to the worksheet
    for row_num, inventory_item in enumerate(inventory_items, 2):
        inventory = inventory_item.inventory
        history_entries = inventory.history_entries.all()

        worksheet.cell(row=row_num, column=1, value=inventory_item.id)
        worksheet.cell(row=row_num, column=2, value=inventory.material.name)
        worksheet.cell(row=row_num, column=3, value=inventory_item.received_quantity)
        worksheet.cell(row=row_num, column=4, value=inventory.description)
        worksheet.cell(row=row_num, column=5, value=inventory.date_created.strftime('%Y-%m-%d %H:%M:%S'))
        worksheet.cell(row=row_num, column=6, value=inventory.reason)
        worksheet.cell(row=row_num, column=7, value=inventory.user.email)

        # Write the history entries to the worksheet
        for history_row_num, history_entry in enumerate(history_entries, 1):
            worksheet.cell(row=row_num + history_row_num, column=8, value=history_entry.reason)
            worksheet.cell(row=row_num + history_row_num, column=9,
                           value=history_entry.date_modified.strftime('%Y-%m-%d %H:%M:%S'))

    # Set the column widths
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 15

    # Create a response with the workbook file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventory.xlsx'
    workbook.save(response)

    return response
