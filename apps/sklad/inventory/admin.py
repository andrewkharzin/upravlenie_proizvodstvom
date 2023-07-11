from django.contrib import admin
from django.http import HttpResponse
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from openpyxl import Workbook
from django.urls import path
from decimal import Decimal
from django.urls import reverse
from django.utils.text import slugify
from django.utils.safestring import mark_safe
from django.utils import timezone
from datetime import datetime
from .admin_site import admin_site

from ..inventory.models.inventory import Inventory, InventoryItem, InventoryHistory


class InventoryResources(resources.ModelResource):
    class Meta:
        model = Inventory
        fields = ('material', 'date_created', 'date_modified')
        import_id_fields = ('date_created',)

    def get_material_info(self, obj):
        return obj.material_info()

    def get_item_count(self, obj):
        return obj.items.count()

    get_material_info.short_description = 'Материал'
    get_item_count.short_description = 'Item Count'



class InventoryItemResources(resources.ModelResource):
    material_info = fields.Field(attribute='material_info', column_name='Material Info')
    total_recieved_cost = fields.Field(attribute='get_total_recieved_cost', column_name='Total Cost')
    

    class Meta:
        model = Inventory
        fields = ('material', 'date_created', 'date_modified', 'material_info', 'get_inv_total_count', 'total_recieved_cost')
        import_id_fields = ('date_created',)

    def get_material_info(self, obj):
        return obj.inventory.material_info()
    
    def get_inv_total_count(self, obj):
        return obj.received_quantity.count()
    

    get_material_info.short_description = 'Material Info'


class InventoryHistoryResources(resources.ModelResource):
    class Meta:
        model = InventoryHistory
        fields = ('inventory__material', 'date_modified', 'reason')
        import_id_fields = ('inventory__material',)

    def get_material_info(self, obj):
        return obj.inventory.material_info()

    get_material_info.short_description = 'Material Info'


class InventoryItemInline(admin.TabularInline):
    model = InventoryItem
    extra = 1


class InventoryAdmin(ImportExportModelAdmin):
    resource_class = InventoryResources
    inlines = [InventoryItemInline]
    list_display = ['inv_number', 'material', 'date_created', 'date_modified', 'material_info', 'get_total_received_quantity', 'get_total_recieved_cost',  'reason', 'user']
    change_list_template = 'admin/inventory_change_list.html'

    raw_id_fields = ['material', ]


    def generate_report_button(self, obj):
        return mark_safe(f'<a href="/admin/inventory/inventory/{obj.id}/generate-report/">Generate Report</a>')
    generate_report_button.short_description = 'Actions'
    generate_report_button.allow_tags = True

    def get_material_info(self, obj):
        return obj.material_info()

    # Получение общего количества полученнго материала
    def get_item_count(self, obj):
        return obj.items.count()
    get_material_info.short_description = 'Материал'

    def get_total_received_quantity(self, obj):
        return sum(item.received_quantity for item in obj.items.all())

    get_total_received_quantity.short_description = 'Получено:'


    #Получение стоимости общего количества полученнго материала
    def get_total_recieved_cost(self, obj):
        total_cost = sum(item.received_quantity for item in obj.items.all()) * obj.material.purchase_price
        # return sum(item.received_quantity for item in obj.items.all()) * obj.material.purchase_price 
        return f"{total_cost} Рублей"

    get_total_recieved_cost.short_description = 'Оценочная стоимость'


    # def calculate_total_cost(self, obj):
    #     total_cost = 0
    #     for item in obj.inventoryitem_set.all():
    #         total_cost += item.material.purchase_price * item.quantity
    #     return total_cost

    # calculate_total_cost.short_description = 'Оценочная стоимость по закупке'

    def generate_report(self, request, inventory_id):
        inventory = Inventory.objects.get(id=inventory_id)
        items = inventory.items.all()

        workbook = Workbook()
        sheet = workbook.active

        # Set headers
        sheet.cell(row=1, column=1, value='Наименование')
        sheet.cell(row=1, column=2, value='Описание')
        sheet.cell(row=1, column=3, value='Дата формирования')
        sheet.cell(row=1, column=4, value='Дата изменений')
        sheet.cell(row=1, column=5, value='Причина')

        row = 2
        for item in items:
            sheet.cell(row=row, column=1, value=item.inventory.name)
            sheet.cell(row=row, column=2, value=item.inventory.description)
            sheet.cell(row=row, column=3, value=item.inventory.date_created)
            sheet.cell(row=row, column=4, value=item.inventory.date_modified)
            sheet.cell(row=row, column=5, value=item.inventory.reason)

            row += 1

        # Convert datetimes to timezone-naive form
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.value = cell.value.replace(tzinfo=None)

        filename = slugify(inventory.name) + '_report.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}

        pdf_report_url = reverse('admin:generate-report', args=[object_id])
        extra_context['pdf_report_url'] = pdf_report_url

        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<path:inventory_id>/generate-report/',
                self.admin_site.admin_view(self.generate_report),
                name='generate-report'
            ),
        ]
        return custom_urls + urls


class InventoryItemAdmin(ImportExportModelAdmin):
    resource_class = InventoryItemResources


class InventoryHistoryAdmin(ImportExportModelAdmin):
    resource_class = InventoryHistoryResources


admin.site.register(Inventory, InventoryAdmin)
admin.site.register(InventoryItem, InventoryItemAdmin)
admin.site.register(InventoryHistory, InventoryHistoryAdmin)

admin_site.register(Inventory, InventoryAdmin)
admin_site.register(InventoryItem, InventoryItemAdmin)
admin_site.register(InventoryHistory, InventoryHistoryAdmin)
