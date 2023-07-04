from django.contrib import admin
from django.http import HttpResponse
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from openpyxl import Workbook
from django.urls import path
from django.urls import reverse
from django.utils.text import slugify
from django.utils.safestring import mark_safe
from django.utils import timezone
from datetime import datetime
from .admin_site import admin_site

from ..inventory.models.inventory import Inventory, InventoryItem, InventoryHistory


class InventoryResources(resources.ModelResource):
    class Meta:
        model = Inventory
        fields = ('material', 
                  'date_created', 
                  'date_modified',  
                  )
        import_id_fields = ('date_created', )   

    def get_material_info(self, obj):
        return obj.material_info()

    get_material_info.short_description = 'Material Info'

class InventoryItemInline(admin.TabularInline):
    model = InventoryItem
    extra = 1


class InventoryAdmin(ImportExportModelAdmin):
    resource_class = InventoryResources
    inlines = [InventoryItemInline]
    list_display = ['get_material_info', 'date_created', 'date_modified', 'reason', 'user']
    change_list_template = 'admin/inventory_change_list.html'

    def generate_report_button(self, obj):
        return mark_safe(f'<a href="/admin/inventory/inventory/{obj.id}/generate-report/">Generate Report</a>')
    generate_report_button.short_description = 'Actions'
    generate_report_button.allow_tags = True


    def get_material_info(self, obj):
        return obj.material_info()

    get_material_info.short_description = 'Material Info'

    def generate_report(self, request, inventory_id):
        inventory = Inventory.objects.get(id=inventory_id)
        items = inventory.items.all()

        workbook = Workbook()
        sheet = workbook.active

        # sheet['A1'] = 'Наименование'
        # sheet['B1'] = 'Описание'
        # sheet['F1'] = 'Дата формирования'
        # sheet['D1'] = 'Дата изменеений'
        # sheet['E1'] = 'Причина'

        row = 2
        for item in items:
            sheet.cell(row=2, column=8, value=item.inventory.name)
            sheet.cell(row=4, column=2, value=item.inventory.description)
            sheet.cell(row=5, column=3, value=item.inventory.date_created)
            sheet.cell(row=6, column=4, value=item.inventory.date_modified)
            sheet.cell(row=12, column=5, value=item.inventory.reason)

            row += 1
        # Convert datetimes to timezone-naive form
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.value = cell.value.replace(tzinfo=None)

        filename = slugify(inventory.name) + '_report.xlsx'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)

        return response
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}

        pdf_report_url = reverse('admin:generate-report', args=[object_id])
        extra_context['pdf_report_url'] = pdf_report_url

        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<path:inventory_id>/generate-report/',
                self.admin_site.admin_view(self.generate_report),
                name='generate-report'
            ),
        ]
        return custom_urls + urls
    


@admin.register(InventoryHistory)
class InventoryHistoryAdmin(admin.ModelAdmin):
    list_display = ('inventory', 'date_modified', 'reason')
    list_filter = ('reason',)
    search_fields = ('inventory__name', 'reason')



admin.site.register(Inventory, InventoryAdmin)
admin.site.register(InventoryItem)

admin_site.register(Inventory)
admin_site.register(InventoryItem)
admin_site.register(InventoryHistory)