from django.contrib import admin
from import_export import resources, fields
from django.http import HttpResponse
from import_export.admin import ImportExportModelAdmin
from mptt.admin import MPTTModelAdmin
from django_mptt_admin.admin import DjangoMpttAdmin
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.sklad.order.models.service_class import Service


class ServiceResources(resources.ModelResource):
    class Meta:
        model = Service
        fields = ('parent', 'name', 'initial_cost',)


class ServiceAdmin(ImportExportModelAdmin, MPTTModelAdmin):
    resource_class = ServiceResources
    list_display = ('name', 'initial_cost', 'price_for_work', 'icon_preview',)
    search_fields = ('name',)
    list_filter = ('parent',)
    mptt_indent_field = 'name'

    def icon_preview(self, obj):
        if obj.icon:
            return mark_safe('<img src="{}" width="50">'.format(obj.icon.url))
        return None
    icon_preview.short_description = 'Icon Preview'


    def generate_report(self, request, queryset):
        workbook = Workbook()
        sheet = workbook.active

        # Set the column headers
        headers = ['Наименование', 'Отпускная базовая цена', 'Учетная стоимость работ']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet['{}1'.format(col_letter)]
            cell.value = header

        # Apply formatting to the parent cell
        bold_font = Font(bold=False)
        fill = PatternFill(start_color="c6def7", end_color="c6def7", fill_type="solid")

        # Populate the data rows
        row_num = 3
        for obj in queryset:
            row = sheet.row_dimensions[row_num]

            sheet.cell(row=row_num, column=1, value=obj.name)
            sheet.cell(row=row_num, column=2, value=obj.initial_cost)
            sheet.cell(row=row_num, column=3, value=obj.price_for_work)

            # Check if the object has a parent
            if obj.parent_id:
                parent_cell = sheet.cell(row=row_num, column=1)
                parent_cell.font = bold_font
                parent_cell.fill = fill

            row_num += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=service_report.xlsx'
        workbook.save(response)

        return response

    generate_report.short_description = 'Generate Report'

    actions = [generate_report]