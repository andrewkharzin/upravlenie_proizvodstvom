from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta
from django.http import HttpResponse


def generate_work_schedule(modeladmin, request, queryset):
    # Create a new instance of the Excel workbook
    workbook = Workbook()
    sheet = workbook.active

    # Define the date range for the schedule
    start_date = min(queryset, key=lambda x: x.start_date).start_date
    end_date = max(queryset, key=lambda x: x.end_date).end_date
    total_days = (end_date - start_date).days + 1

    # Formatting styles
    title_font = Font(name='Arial', size=14, bold=True, color='000000')
    header_font = Font(name='Arial', size=12, bold=True, color='000000')
    cell_font = Font(name='Arial', size=11)
    header_fill = PatternFill(fill_type='solid', fgColor='DDEBF7')
    work_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
    weekend_fill = PatternFill(fill_type='solid', fgColor='FCE5CD')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Fill the title
    sheet.merge_cells('A1:C1')
    sheet['A1'] = 'Work Schedule'
    sheet['A1'].font = title_font
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 30

    # Fill the column headers with weekdays and dates
    weekdays = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    column = 4  # Starting column for data
    row = 3  # Starting row for data
    for day in range(total_days):
        date = start_date + timedelta(days=day)
        column_letter = get_column_letter(column)
        cell = sheet[column_letter + str(row)]

        # Move to the next row if it's a Monday
        if date.weekday() == 0:
            row += 1

        sheet[column_letter + str(row)] = weekdays[date.weekday()] + '\n' + date.strftime('%d.%m')
        sheet[column_letter + str(row)].font = header_font
        sheet[column_letter + str(row)].alignment = Alignment(horizontal='center', vertical='center')
        sheet[column_letter + str(row)].fill = header_fill
        sheet.column_dimensions[column_letter].width = 12
        column += 1

    # Fill the data for work schedule
    row = 4  # Starting row for data (skip the header row)
    for workshift in queryset:
        sheet['A' + str(row)] = workshift.employee.name
        sheet['B' + str(row)] = workshift.employee.position
        sheet['A' + str(row)].font = cell_font
        sheet['B' + str(row)].font = cell_font
        sheet['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
        sheet['B' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A' + str(row)].fill = work_fill
        sheet['B' + str(row)].fill = work_fill
        sheet['A' + str(row)].border = border
        sheet['B' + str(row)].border = border

        column = 4  # Starting column for data
        for day in range(total_days):
            date = start_date + timedelta(days=day)
            column_letter = get_column_letter(column)

            # Move to the next row if it's a Monday
            if date.weekday() == 0:
                row += 1

            cell = sheet[column_letter + str(row)]

            if workshift.start_date <= date <= workshift.end_date:
                # Fill the cell for a work day
                cell.value = 'Р'
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = work_fill
                cell.border = border
            elif date.weekday() in [5, 6]:  # Saturday (5) and Sunday (6)
                # Fill the cell for a weekend day
                cell.value = 'ВХ'
                cell.font = cell_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = weekend_fill
                cell.border = border

            column += 1

        row += 1

    # Generate the filename for downloading
    filename = 'work_schedule.xlsx'

    # Create an HTTP response for downloading the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename

    # Save the workbook to the HTTP response
    workbook.save(response)
    return response
